import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _style_cell(cell, bold=False, fill_color=None, align=None):
    if bold: cell.font = Font(bold=True)
    if fill_color: cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if align: cell.alignment = Alignment(horizontal=align)
    else: cell.alignment = Alignment(horizontal="left") # Default to left
    
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def generate_xlsx(payload: Dict[str, Any], out_path: Path, history: Optional[List[Dict[str, Any]]] = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lieferplan"

    # 1. Summary Section Header
    curr_row = 1
    cell = ws.cell(row=curr_row, column=1, value="SOUHRN INFORMACÍ")
    _style_cell(cell, bold=True, fill_color="D9D9D9")
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
    curr_row += 1

    labels = [
        ("Číslo Lieferplánu", payload.get("scheduling_agreement_no")),
        ("Změnovka (Release Nr.)", payload.get("release_nr")),
        ("-", "-"),
        ("Dodací Adresa", payload.get("receiving_factory")),
        ("Rampa", payload.get("warehouse_rampe")),
        ("-", "-"),
        ("Číslo výrobku", payload.get("material_no")),
        ("Balení", payload.get("pal_typ")),
        ("Množství", payload.get("volume_value", 0)),
    ]

    for label, val in labels:
        if label == "-":
            curr_row += 1
            continue
        
        c1 = ws.cell(row=curr_row, column=1, value=label)
        c2 = ws.cell(row=curr_row, column=2, value=val or "Chybí")
        _style_cell(c1, bold=True, fill_color="F2F2F2")
        _style_cell(c2, align="left")
        
        if val is None:
            c2.font = Font(color="D32F2F", italic=True)

        if label == "Množství" and val is not None:
            c2.number_format = '#,##0'
            unit = payload.get("volume_unit")
            if unit:
                c1.value = f"{label} (v {unit})"
        curr_row += 1

    curr_row += 2
    # 2. Schedule Section Header
    cell = ws.cell(row=curr_row, column=1, value="HARMONOGRAM DODÁVEK")
    _style_cell(cell, bold=True, fill_color="D9D9D9")
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
    curr_row += 1

    headers = ["#", "Termín dodání", "Objednané množství", "Změna", "Dní do dodání"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=curr_row, column=i, value=h)
        _style_cell(cell, bold=True, fill_color="D9EAD3", align="center")
    
    curr_row += 1
    today = datetime.now().date()
    
    for idx, line in enumerate(payload.get("lines", []), 1):
        dd_raw = line.get("delivery_date") # YYYY-MM-DD
        display_date = dd_raw
        days_to = ""
        row_fill = None
        is_bold = True
        
        try:
            dd = datetime.strptime(dd_raw, "%Y-%m-%d").date()
            display_date = dd.strftime("%d.%m.%Y")
            diff = (dd - today).days
            days_to = diff
            
            modification = line.get("modification")
            
            if diff < 0:
                row_fill = "F9F9F9" # Lightest grey for past items
                is_bold = False # Unbold past items
            elif diff < 45 and modification and modification != 0:
                row_fill = "F4CCCC" # Red for urgent modifications
        except Exception: 
            pass

        row_data = [
            idx,
            display_date,
            line.get("order_quantity"),
            line.get("modification"),
            days_to
        ]
        for i, val in enumerate(row_data, 1):
            cell = ws.cell(row=curr_row, column=i, value=val)
            _style_cell(cell, bold=is_bold and i==2, align="center", fill_color=row_fill)
            
            if i == 3: # Quantity
                cell.number_format = '#,##0'
            if i == 4 and val: # Modification
                cell.number_format = '+#,##0;-#,##0;0'
        curr_row += 1

    # Safe autosize columns
    for col_idx in range(1, 6):
        max_w = 0
        column_letter = get_column_letter(col_idx)
        for row in range(1, curr_row):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                lines = str(cell.value).split('\n')
                max_w = max(max_w, max(len(ln) for ln in lines))
        ws.column_dimensions[column_letter].width = min(40, max_w + 5)

    # 3. History Sheet
    h_ws = wb.create_sheet("Historie verzí")
    h_headers = ["Datum nahrání", "Změnovka", "Číslo výrobku", "Složka"]
    for i, h in enumerate(h_headers, 1):
        cell = h_ws.cell(row=1, column=i, value=h)
        _style_cell(cell, bold=True, fill_color="CFE2F3", align="center")
    
    if history:
        for r_idx, entry in enumerate(history, 2):
            h_ws.cell(row=r_idx, column=1, value=entry.get("uploaded_at"))
            h_ws.cell(row=r_idx, column=2, value=entry.get("release_nr"))
            h_ws.cell(row=r_idx, column=3, value=entry.get("material_no"))
            h_ws.cell(row=r_idx, column=4, value=entry.get("plan_id"))
            for i in range(1, 5):
                _style_cell(h_ws.cell(row=r_idx, column=i), align="left")
    
    for col_idx in range(1, 5):
        max_w = 0
        column_letter = get_column_letter(col_idx)
        for row in range(1, (len(history) if history else 0) + 2):
            cell = h_ws.cell(row=row, column=col_idx)
            if cell.value:
                max_w = max(max_w, len(str(cell.value)))
        h_ws.column_dimensions[column_letter].width = max_w + 5

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _register_unicode_font():
    """Register a Unicode-capable TTF font for ReportLab (supports Czech diacritics)."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    # Try Arial (always present on Windows), fall back to DejaVuSans
    candidates = [
        ("C:/Windows/Fonts/arial.ttf",   "C:/Windows/Fonts/arialbd.ttf"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ]
    for regular, bold in candidates:
        if os.path.exists(regular) and os.path.exists(bold):
            pdfmetrics.registerFont(TTFont("UniFont",     regular))
            pdfmetrics.registerFont(TTFont("UniFont-Bold", bold))
            return "UniFont", "UniFont-Bold"
    # Last-resort: built-in (diacritics will still be broken, but won't crash)
    return "Helvetica", "Helvetica-Bold"

def generate_pdf(payload: Dict[str, Any], out_path: Path, history: Optional[List[Dict[str, Any]]] = None) -> None:
    """
    Generate a PDF export of the Lieferplan using reportlab.
    Mirrors the structure of generate_xlsx: summary block + schedule table + optional version history.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    out_path.parent.mkdir(parents=True, exist_ok=True)

    FONT, FONT_BOLD = _register_unicode_font()

    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=14, spaceAfter=6, textColor=colors.HexColor("#1a1a2e"), fontName=FONT_BOLD)
    section_style = ParagraphStyle("section", parent=styles["Heading2"], fontSize=10, spaceAfter=4, textColor=colors.HexColor("#1976d2"), spaceBefore=8, fontName=FONT_BOLD)
    normal_style = ParagraphStyle("normal", parent=styles["Normal"], fontSize=9, fontName=FONT)
    small_style = ParagraphStyle("small", parent=styles["Normal"], fontSize=7.5, textColor=colors.grey, fontName=FONT)

    # Color palette matching the app
    C_HEADER_BG  = colors.HexColor("#1976d2")
    C_HEADER_TXT = colors.white
    C_ROW_ALT    = colors.HexColor("#f0f4ff")
    C_ROW_WHITE  = colors.white
    C_URGENT     = colors.HexColor("#f4cccc")
    C_PAST       = colors.HexColor("#f5f5f5")
    C_SECTION_BG = colors.HexColor("#e3f2fd")
    C_BORDER     = colors.HexColor("#c8d6e5")
    C_LABEL_BG   = colors.HexColor("#f8fafc")

    story = []

    # ── Title ────────────────────────────────────────────────────────────────
    sa_no = payload.get("scheduling_agreement_no") or "—"
    rel_nr = payload.get("release_nr") or "—"
    story.append(Paragraph(f"Lieferplan: {sa_no}", title_style))
    story.append(Paragraph(f"Změnovka (Release): {rel_nr}", small_style))
    story.append(Spacer(1, 0.3 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=C_BORDER))
    story.append(Spacer(1, 0.3 * cm))

    # ── Summary ───────────────────────────────────────────────────────────────
    story.append(Paragraph("Souhrn informací", section_style))

    def _val(v, fallback="Chybí"):
        return str(v) if v not in (None, "", []) else fallback

    vol = f"{payload.get('volume_value', '')} {payload.get('volume_unit') or ''}".strip() or "Chybí"

    summary_data = [
        ["Číslo Lieferplánu", _val(payload.get("scheduling_agreement_no"))],
        ["Změnovka (Release Nr.)", _val(rel_nr)],
        ["Dodací Adresa", _val(payload.get("receiving_factory"))],
        ["Rampa", _val(payload.get("warehouse_rampe"))],
        ["Číslo výrobku", _val(payload.get("material_no"))],
        ["Balení", _val(payload.get("pal_typ"))],
        ["Množství", vol],
    ]

    sum_table = Table(summary_data, colWidths=[5.5 * cm, None])
    sum_ts = TableStyle([
        ("BACKGROUND",   (0, 0), (0, -1), C_LABEL_BG),
        ("FONTNAME",     (0, 0), (-1, -1), FONT),
        ("FONTNAME",     (0, 0), (0, -1), FONT_BOLD),
        ("FONTSIZE",     (0, 0), (-1, -1), 9),
        ("TEXTCOLOR",    (0, 0), (0, -1), colors.HexColor("#334155")),
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("GRID",         (0, 0), (-1, -1), 0.5, C_BORDER),
        ("ROWBACKGROUNDS", (1, 0), (1, -1), [C_ROW_WHITE, C_ROW_ALT]),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])
    sum_table.setStyle(sum_ts)
    story.append(sum_table)
    story.append(Spacer(1, 0.5 * cm))

    # ── Schedule Table ────────────────────────────────────────────────────────
    story.append(Paragraph("Harmonogram dodávek", section_style))

    today = datetime.now().date()
    headers = ["#", "Termín dodání", "Objednané množství", "Změna", "Dní do dodání"]
    sched_data = [headers]

    for idx, line in enumerate(payload.get("lines", []), 1):
        dd_raw = line.get("delivery_date", "")
        display_date = dd_raw
        days_to = ""
        try:
            dd = datetime.strptime(dd_raw, "%Y-%m-%d").date()
            display_date = dd.strftime("%d.%m.%Y")
            days_to = (dd - today).days
        except Exception:
            pass

        mod = line.get("modification")
        mod_str = ""
        if mod is not None:
            mod_str = f"+{mod}" if mod > 0 else str(mod)

        qty = line.get("order_quantity", "")
        qty_str = f"{qty:,}".replace(",", " ") if isinstance(qty, int) else str(qty)

        sched_data.append([str(idx), display_date, qty_str, mod_str, str(days_to) if days_to != "" else ""])

    col_widths = [1.0 * cm, 3.2 * cm, 4.2 * cm, 2.5 * cm, 3.5 * cm]
    sched_table = Table(sched_data, colWidths=col_widths)

    sched_ts = TableStyle([
        # Header row
        ("BACKGROUND",   (0, 0), (-1, 0), C_HEADER_BG),
        ("TEXTCOLOR",    (0, 0), (-1, 0), C_HEADER_TXT),
        ("FONTNAME",     (0, 0), (-1, -1), FONT),
        ("FONTNAME",     (0, 0), (-1, 0), FONT_BOLD),
        ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
        ("FONTSIZE",     (0, 0), (-1, -1), 9),
        ("GRID",         (0, 0), (-1, -1), 0.5, C_BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_ROW_WHITE, C_ROW_ALT]),
        ("ALIGN",        (0, 1), (0, -1), "CENTER"),   # #
        ("ALIGN",        (1, 1), (1, -1), "CENTER"),   # date
        ("ALIGN",        (2, 1), (2, -1), "RIGHT"),    # qty
        ("ALIGN",        (3, 1), (3, -1), "CENTER"),   # mod
        ("ALIGN",        (4, 1), (4, -1), "CENTER"),   # days
        ("TOPPADDING",   (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ])

    # Row-level highlighting
    for row_idx, line in enumerate(payload.get("lines", []), 1):
        dd_raw = line.get("delivery_date", "")
        try:
            dd = datetime.strptime(dd_raw, "%Y-%m-%d").date()
            diff = (dd - today).days
            mod = line.get("modification")
            if diff < 0:
                sched_ts.add("BACKGROUND", (0, row_idx), (-1, row_idx), C_PAST)
                sched_ts.add("TEXTCOLOR",  (0, row_idx), (-1, row_idx), colors.grey)
            elif diff < 45 and mod and mod != 0:
                sched_ts.add("BACKGROUND", (0, row_idx), (-1, row_idx), C_URGENT)
        except Exception:
            pass

    sched_table.setStyle(sched_ts)
    story.append(sched_table)

    # ── Version History ───────────────────────────────────────────────────────
    if history:
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph("Historie verzí", section_style))

        h_data = [["Datum nahrání", "Změnovka", "Číslo výrobku"]]
        for entry in history:
            h_data.append([
                entry.get("uploaded_at", ""),
                entry.get("release_nr", ""),
                entry.get("material_no", ""),
            ])

        h_table = Table(h_data, colWidths=[5 * cm, 3 * cm, None])
        h_ts = TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), C_SECTION_BG),
            ("FONTNAME",     (0, 0), (-1, -1), FONT),
            ("FONTNAME",     (0, 0), (-1, 0), FONT_BOLD),
            ("FONTSIZE",     (0, 0), (-1, -1), 8.5),
            ("GRID",         (0, 0), (-1, -1), 0.5, C_BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_ROW_WHITE, C_ROW_ALT]),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ])
        h_table.setStyle(h_ts)
        story.append(h_table)

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.5 * cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER))
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    story.append(Paragraph(f"Vygenerováno: {generated_at}", small_style))

    doc.build(story)


def main() -> None:

    parser = argparse.ArgumentParser(description="Generate Lieferplan XLSX")
    parser.add_argument("input_json", type=Path, help="Extracted JSON file")
    parser.add_argument("--out", type=Path, required=True, help="Output XLSX file")
    parser.add_argument("--changelog", type=Path, help="Optional changelog JSON file")
    args = parser.parse_args()
    with args.input_json.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    changelog_data = None
    if args.changelog and args.changelog.exists():
        with args.changelog.open("r", encoding="utf-8") as handle:
            changelog_data = json.load(handle)
    generate_xlsx(payload, args.out, history=changelog_data)

if __name__ == "__main__":
    main()
