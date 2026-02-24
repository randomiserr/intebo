import openpyxl
from pathlib import Path
from typing import Dict

def parse_inventory_xlsx(file_path: Path) -> Dict[str, float]:
    """
    Parses a POHODA inventory Excel file and extracts warehouse quantities per material.
    
    The file structure has merged cells in columns T:U for the 'Množství' column.
    We return a dictionary mapping material_code -> quantity.
    """
    inventory_data = {}
    
    # We must use data_only=True to get values instead of formulas,
    # and we cannot use read_only=True because we need to read merged cells properties.
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Step 1: Unmerge T:U cells so we can read the value from column T (col 20)
    # The merged cells in this specific area cause the value to sit in the top-left cell of the merge.
    for mc in list(ws.merged_cells.ranges):
        if mc.min_col == 20 and mc.max_col == 21:
            try:
                ws.unmerge_cells(str(mc))
            except KeyError:
                # Sometimes unmerge fails if it was already unmerged or invalid
                pass
                
    # Step 2: Read Data rows (start around row 5)
    # Kod is column 2 (B), Mnozstvi is column 20 (T) after unmerging
    for i in range(5, ws.max_row + 1):
        kod_val = ws.cell(row=i, column=2).value
        qty_val = ws.cell(row=i, column=20).value
        
        if kod_val is not None:
            kod_str = str(kod_val).strip()
            if kod_str:
                try:
                    qty = float(qty_val) if qty_val is not None else 0.0
                except (ValueError, TypeError):
                    qty = 0.0
                
                # If a code appears multiple times, sum it up (though it should be unique usually)
                if kod_str in inventory_data:
                    inventory_data[kod_str] += qty
                else:
                    inventory_data[kod_str] = qty
                    
    wb.close()
    return inventory_data
